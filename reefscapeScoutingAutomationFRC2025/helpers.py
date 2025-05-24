import openpyxl as xl
import time
import os

def team_assign(wb, reference_sheet, duplicate_sheet):
    reference_team = wb[reference_sheet]
    duplicate_team = wb[duplicate_sheet]

    for row in range(2, reference_team.max_row + 1):
        # Variables
        old_sheet_cell = reference_team.cell(row, 1)
        new_sheet_cell = duplicate_team.cell(row, 1)

        # Assigning values to respective cells
        new_sheet_cell.value = old_sheet_cell.value
        print(f"Row {row}: Copying team number {old_sheet_cell.value}")  # check what's being copied

def is_file_open(filepath):
    """Returns True if the file is currently open/locked."""
    try:
        with open(filepath, 'r+'):
            return False
    except IOError:
        return True