from safety_variable_assign import assign_value
import openpyxl as xl

filepath = '../data/Reefscape Scouting Sheet.xlsx'

# Loads the workbook
scouting_wb = xl.load_workbook(filepath)

# Scouting Sheet Controller
team_number = 1
match = 1

# Auto Data
auto_leave = 1  # 1 if yes, 0 if no
auto_lvl_4 = 1
auto_lvl_3 = 1
auto_lvl_2 = 1
auto_lvl_1 = 1
auto_processor = 1
auto_barge_net = 1

# TeleOp Data
teleop_lvl_4 = 1
teleop_lvl_3 = 1
teleop_lvl_2 = 1
teleop_lvl_1 = 1
teleop_processor = 1
teleop_barge_net = 1

# End Game
park = 1
shallow_climb = 1
deep_climb = 1

unsorted_sheet_auto = scouting_wb['Auto - Unsorted Data']
unsorted_sheet_teleop = scouting_wb['TeleOp - Unsorted Data']

# Team Number and Match
unsorted_sheet_auto.cell(unsorted_sheet_auto.max_row + 1, 1).value = team_number
unsorted_sheet_teleop.cell(unsorted_sheet_teleop.max_row + 1, 1).value = team_number

assign_value(unsorted_sheet_auto, 9, match)
assign_value(unsorted_sheet_teleop, 11, match)

# Auto Assign
assign_value(unsorted_sheet_auto, 2, auto_lvl_1)
assign_value(unsorted_sheet_auto, 3, auto_lvl_2)
assign_value(unsorted_sheet_auto, 4, auto_lvl_3)
assign_value(unsorted_sheet_auto, 5, auto_lvl_4)
assign_value(unsorted_sheet_auto, 6, auto_processor)
assign_value(unsorted_sheet_auto, 7, auto_barge_net)
assign_value(unsorted_sheet_auto, 8, auto_leave)
assign_value(unsorted_sheet_auto, 10, 0)

# TeleOp Assign
assign_value(unsorted_sheet_teleop, 2, teleop_lvl_1)
assign_value(unsorted_sheet_teleop, 3, teleop_lvl_2)
assign_value(unsorted_sheet_teleop, 4, teleop_lvl_3)
assign_value(unsorted_sheet_teleop, 5, teleop_lvl_4)
assign_value(unsorted_sheet_teleop, 6, teleop_processor)
assign_value(unsorted_sheet_teleop, 7, teleop_barge_net)
assign_value(unsorted_sheet_teleop, 8, park)
assign_value(unsorted_sheet_teleop, 9, shallow_climb)
assign_value(unsorted_sheet_teleop, 10, deep_climb)
assign_value(unsorted_sheet_teleop, 12, 0)

scouting_wb.save(filepath)