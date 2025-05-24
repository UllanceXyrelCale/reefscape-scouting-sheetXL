import openpyxl as xl
filepath = '../data/Reefscape Scouting Sheet.xlsx'
scouting_wb = xl.load_workbook(filepath)


def teleop_data_sorter(wb, filename, unsorted_sheet, sorted_sheet):
    unsorted_sheet = wb[unsorted_sheet]
    sorted_sheet = wb[sorted_sheet]

    for unsorted_cell_row in range(2, unsorted_sheet.max_row + 1):
        if unsorted_sheet.cell(unsorted_cell_row, 12).value == 0:  # Check if the team's data hasn't been read

            for sorted_cell_row in range(2, unsorted_sheet.max_row + 1):
                if unsorted_sheet.cell(unsorted_cell_row, 1).value == sorted_sheet.cell(sorted_cell_row, 1).value:# If yes, check the team number of that data that matches on the other workbook and transfer the data there

                    for transfer_column in range(2, unsorted_sheet.max_column):
                        try:
                            new_data = unsorted_sheet.cell(unsorted_cell_row, transfer_column).value + sorted_sheet.cell(sorted_cell_row, transfer_column).value
                            sorted_sheet.cell(sorted_cell_row, transfer_column).value = new_data
                            wb.save(filename)
                        except TypeError:
                            print('error detected')
                    unsorted_sheet.cell(unsorted_cell_row, 12).value = 1
                    wb.save(filename)
                    break
                elif sorted_cell_row == unsorted_sheet.max_row:
                    control = 0
                    for transfer_column in range(1, unsorted_sheet.max_column):

                        try:
                            sorted_sheet.cell(sorted_sheet.max_row + 1 - control, transfer_column).value = unsorted_sheet.cell(unsorted_cell_row, transfer_column).value
                            control = control + 1
                            if control > 1:
                                control = 1
                            print(sorted_sheet.max_row + 1 - control)
                            wb.save(filename)
                        except TypeError:
                            print('error detected')
                    unsorted_sheet.cell(unsorted_cell_row, 12).value = 1
                    wb.save(filename)
                    break

def auto_data_sorter(wb, filename, unsorted_sheet, sorted_sheet):
    unsorted_sheet = wb[unsorted_sheet]
    sorted_sheet = wb[sorted_sheet]

    for unsorted_cell_row in range(2, unsorted_sheet.max_row + 1):
        if unsorted_sheet.cell(unsorted_cell_row, 10).value == 0:  # Check if the team's data hasn't been read

            for sorted_cell_row in range(2, unsorted_sheet.max_row + 1):
                if unsorted_sheet.cell(unsorted_cell_row, 1).value == sorted_sheet.cell(sorted_cell_row, 1).value:# If yes, check the team number of that data that matches on the other workbook and transfer the data there

                    for transfer_column in range(2, unsorted_sheet.max_column):
                        try:
                            new_data = unsorted_sheet.cell(unsorted_cell_row, transfer_column).value + sorted_sheet.cell(sorted_cell_row, transfer_column).value
                            sorted_sheet.cell(sorted_cell_row, transfer_column).value = new_data
                            wb.save(filename)
                        except TypeError:
                            print('error detected')
                    unsorted_sheet.cell(unsorted_cell_row, 10).value = 1
                    wb.save(filename)
                    break
                elif sorted_cell_row == unsorted_sheet.max_row:
                    control = 0
                    for transfer_column in range(1, unsorted_sheet.max_column):

                        try:
                            sorted_sheet.cell(sorted_sheet.max_row + 1 - control, transfer_column).value = unsorted_sheet.cell(unsorted_cell_row, transfer_column).value
                            control = control + 1
                            if control > 1:
                                control = 1
                            print(sorted_sheet.max_row + 1 - control)
                            wb.save(filename)
                        except TypeError:
                            print('error detected')
                    unsorted_sheet.cell(unsorted_cell_row, 10).value = 1
                    wb.save(filename)
                    break
